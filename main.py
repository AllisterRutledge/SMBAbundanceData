# """Creates an abundance dataframe from occupancy excel files in current directory and then appends df to new or
# existing abundance excel file. Reformats excel file when finished. """
import datetime
import os
import re
import sys
import traceback

import pandas
from openpyxl import load_workbook, styles

SMB_IDS = ['COGA', 'CLING', 'CLRA', 'KIRA', 'PUGA', 'LEBI', 'SORA', 'AMCO', 'PBGR', 'LIMP']
OCC_HEADERS = ['Site', 'Point', 'Bout', 'Date', 'Time', 'Full Point Count (Y/N)', 'Observer', 'Sky',
               'Wind Speed (knots)', 'Temp (C)', 'Noise', 'Water Depth (m)']
ABU_HEADERS = ['Site', 'Point', 'Bout', 'Date', 'Time', 'Full Point Count (Y/N)', 'Observer', 'Sky',
               'Wind', 'Temp', 'Sound', 'Water Depth']


def find_excel_file(keyword, output_word):
    """Checks current directory for .xlsx files containing the keyword. Asks for user input if more than 1 file found.
     Returns filename as a string"""
    excel_files = [file for file in os.listdir() if re.findall(f"{keyword}.*.xlsx", file, re.IGNORECASE)]
    if len(excel_files) == 1:
        return excel_files[0]
    if len(excel_files) > 1:
        print(f"More than 1 possible {output_word} file found:")
        for file in excel_files:
            print(f"{excel_files.index(file) + 1}: {file}")
        user_input = int(input(f"\nEnter index number of which {output_word} file to use (1-{len(excel_files)}): "))
        return excel_files[user_input - 1]
    raise FileNotFoundError(f"No {output_word} file could be found based on search expression '{keyword}.*.xlsx'")


def find_excel_sheet(excel_file):
    """Returns sheet name as string from excel file. if more than 1 sheet is found, asks for user input."""
    book = load_workbook(excel_file)
    if len(book.sheetnames) > 1:
        print(f"More than 1 sheet found in excel file:")
        for sheet in book.sheetnames:
            print(f"{book.sheetnames.index(sheet) + 1}: {sheet}")
        user_input = int(input(f"\nEnter index number of which sheet to use (1-{len(book.sheetnames)}): "))
        return book.sheetnames[user_input - 1]
    return book.sheetnames[0]


def wrong_headers(excel_file, excel_sheet):
    """Returns boolean based on if first row of excel sheet matches ABU_HEADERS with SMD_IDS"""
    book = load_workbook(excel_file)
    sheet = book[excel_sheet]
    correct_headers = ABU_HEADERS + SMB_IDS
    correct_headers.append("Proofed by")
    for header in correct_headers:
        if header == sheet.cell(row=1, column=(correct_headers.index(header) + 1)).value:
            continue
        else:
            print("Adding headers to Abundance file.")
            return True
    return False


def user_continue(excel_file_name, excel_sheet_name, output_word, additional_options=False):
    """Asks user if they want to continue with selected file for {output_word} data file. Returns boolean."""
    print(f"\nLoaded FILE: '{excel_file_name}' & SHEET: '{excel_sheet_name}' as {output_word} data file.\n")
    if additional_options:
        user_input = input("Hit 'any key' and then 'Enter' to cancel and bring up additional options or\n"
                           "Press 'Enter' to continue with this file. ")
    else:
        user_input = input("Hit 'any key' and then 'Enter' to cancel. Press 'Enter' to continue with this file. ")
    if user_input != "":
        return False
    return True


# use try/except statement to catch and print any error before the .exe console window closes.
try:

    # create abundance df from occupancy excel file.
    print("Loading Occupancy data file...")
    excel_occupancy_file = find_excel_file("occupanc", "Occupancy")
    excel_occupancy_sheet = find_excel_sheet(excel_occupancy_file)
    smbdf = pandas.read_excel(excel_occupancy_file, sheet_name=excel_occupancy_sheet)
    if not user_continue(excel_occupancy_file, excel_occupancy_sheet, "Occupancy"):
        input("\nUser cancelled program. Hit enter to close")
        sys.exit()
    print("Creating abundance dataframe...")
    smbdf = smbdf[smbdf["Proofed By"].isna()]
    # create list of sites and points
    site_list = [row["Site"] for index, row in smbdf.iterrows()]
    site_list = list(dict.fromkeys(site_list))

    abundance_entry_list = []
    for site in site_list:
        site_df = smbdf[smbdf["Site"] == site]
        while not site_df.empty:
            point = site_df.iloc[0, 1]
            point_df = site_df[site_df["Point"] == point]
            row = point_df[OCC_HEADERS]
            row = row.iloc[[0], :]
            row = row.rename({"Wind Speed (knots)": "Wind", "Temp (C)": "Temp", "Noise": "Sound",
                              "Water Depth (m)": "Water Depth"}, axis="columns", errors="raise")
            for bird in SMB_IDS:
                species_series = point_df["Species Code"]
                abundance = species_series[species_series == bird].count()
                row[bird] = abundance
            row["Proofed by"] = point_df["Proofed By"].iloc[0]
            abundance_entry_list.append(row)
            site_df = site_df[site_df["Point"] != point]
    try:
        abundance_df = pandas.concat(abundance_entry_list)
    except ValueError as msg:
        print(msg)
        input("No new data found in occupancy data file. "
              "All the occupancy data has been already proofed or the the file is empty."
              "\n\nPress enter to close program.")
        sys.exit()

    # output abundance df to existing or new excel file
    print("Loading abundance data file...")
    abundance_file_date = abundance_df["Date"].iat[0].strftime("%m-%d-%Y")

    try:  # Try to update existing file first
        excel_abundance_file = find_excel_file("abundanc", "Abundance")
        excel_abundance_sheet = find_excel_sheet(excel_abundance_file)
        # Ask user if they want to make a new abundance file instead if they decline to continue with file.
        if not user_continue(excel_abundance_file, excel_abundance_sheet, "Abundance", True):
            make_file = input("\nUser declined using file. Would you like to make a new Abundance file instead? "
                              "\nPress 'f' and then enter to make a new file,\npress 's' to add a new sheet to file "
                              f"'{excel_abundance_file}', \nor press 'enter' to cancel: ").lower()
            if make_file == "f":
                raise FileNotFoundError
            elif make_file == "s":
                wb = load_workbook(excel_abundance_file)
                if abundance_file_date in wb.sheetnames:  # Check if overwriting sheet in workbook
                    wb.create_sheet(abundance_file_date + f' (new {datetime.datetime.now().strftime("%m-%d %H%M%S")})')
                else:
                    wb.create_sheet(abundance_file_date)
                wb.save(excel_abundance_file)
                excel_abundance_sheet = wb.sheetnames[-1]
            else:
                input("\nUser cancelled program. Hit enter to close")
                sys.exit()
        are_headers_wrong = wrong_headers(excel_abundance_file, excel_abundance_sheet)
        with pandas.ExcelWriter(excel_abundance_file,
                                mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            original_abu_df = pandas.read_excel(excel_abundance_file, sheet_name=excel_abundance_sheet)
            starting_row = 0
            if len(original_abu_df) > 0:
                starting_row = len(original_abu_df) + 1
            abundance_df.to_excel(writer, sheet_name=excel_abundance_sheet, index=False,
                                  startrow=starting_row, header=are_headers_wrong)
        print("Non-proofed Occupancy data added to Abundance file.")
    except FileNotFoundError:  # Create new abundance file:
        print("No file found. Creating abundance file.")
        excel_abundance_sheet = abundance_file_date
        # Try to change name of occupancy file
        word_to_replace = re.findall("Occupanc.", excel_occupancy_file, re.IGNORECASE)
        excel_abundance_file = excel_occupancy_file.replace(word_to_replace[0], "Abundance")
        if excel_abundance_file == excel_occupancy_file:  # if name change didn't work create file name with date
            excel_abundance_file = f"{abundance_file_date} Abundance Data Entry.xlsx"
        # Check if new abundance file name already exists so old file is not overwritten.
        # Will only come up if line 130 equals True and user chooses to make a new file.
        abundance_file_names = [file for file in os.listdir() if re.findall(excel_abundance_file, file)]
        if len(abundance_file_names) > 0:
            excel_abundance_file = excel_abundance_file.replace(
                ".xlsx", f' (new {datetime.datetime.now().strftime("%m-%d %H%M%S")}).xlsx')
        with pandas.ExcelWriter(excel_abundance_file) as writer:
            abundance_df.to_excel(writer, sheet_name=abundance_file_date, index=False)
        print(f"New abundance file created: '{excel_abundance_file}'")

    # Format the excel file
    print("Reformatting abundance file...")
    wb = load_workbook(excel_abundance_file)
    ws = wb[excel_abundance_sheet]
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.border = styles.Border()
    for col in ws["I":"J"]:
        for cell in col:
            cell.number_format = "0.0"
    for cell in ws["D"]:
        cell.number_format = "MM/DD/YYYY"
    for cell in ws["E"][1:]:
        # if cell is time formatted as string then convert to time object, otherwise do nothing.
        if type(cell.value) == str:
            cell.value = datetime.time.fromisoformat(cell.value)
            cell.number_format = "h:mm"
    for cell in ws["W"]:
        cell.fill = styles.PatternFill("solid", fgColor="FFFF00")

    wb.save(excel_abundance_file)
    input("Finished. Press enter to close.")

except Exception as e:
    print(traceback.print_exc())
    input("\nAn error occurred (see above). Press enter to close window.")
